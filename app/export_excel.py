"""Multi-sheet Excel export with formatting for menu comparison reports."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Display column order for the main comparison sheet
COMPARISON_COLUMNS = [
    "platform",
    "online_name",
    "online_category",
    "online_family",
    "pos_name",
    "pos_category",
    "pos_family",
    "pos_id",
    "online_price",
    "pos_price",
    "diff",
    "diff_pct",
    "match_score",
    "category_match_ok",
    "action",
]

HEADER_FILL = PatternFill(start_color="06C167", end_color="06C167", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
WARN_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
BAD_FILL = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
OK_FILL = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")


def _order_columns(df: pd.DataFrame, preferred: list[str]) -> pd.DataFrame:
    cols = [c for c in preferred if c in df.columns]
    extra = [c for c in df.columns if c not in cols]
    return df[cols + extra]


def _format_sheet(ws, df: pd.DataFrame, *, currency_cols: tuple[str, ...] = ()) -> None:
    """Apply headers, column widths, and conditional row colours."""
    if df.empty:
        return

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for idx, col in enumerate(df.columns, start=1):
        letter = get_column_letter(idx)
        max_len = max(len(str(col)), df[col].astype(str).str.len().max() if len(df) else 0)
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 45)

        if col in currency_cols:
            for row in range(2, len(df) + 2):
                ws[f"{letter}{row}"].number_format = "£#,##0.00"
        if col == "diff_pct":
            for row in range(2, len(df) + 2):
                ws[f"{letter}{row}"].number_format = '0.0"%"'

    if "action" in df.columns:
        action_col = list(df.columns).index("action") + 1
        letter = get_column_letter(action_col)
        for row in range(2, len(df) + 2):
            action = str(ws[f"{letter}{row}"].value or "")
            fill = None
            if action == "ok":
                fill = OK_FILL
            elif "unmatched" in action or "category_mismatch" in action or "low_confidence" in action:
                fill = WARN_FILL
            elif action in ("lower_online_or_raise_pos", "raise_online_or_lower_pos"):
                fill = BAD_FILL
            if fill:
                for c in range(1, len(df.columns) + 1):
                    ws[f"{get_column_letter(c)}{row}"].fill = fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def export_report(
    output_path: str | Path,
    comparison: pd.DataFrame,
    unmatched_online: pd.DataFrame,
    unmatched_pos: pd.DataFrame,
    menus_raw: pd.DataFrame,
    pos_raw: pd.DataFrame,
) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if "uber_name" in comparison.columns and "online_name" not in comparison.columns:
        unmatched_online = unmatched_online.rename(
            columns={"uber_name": "online_name"}, errors="ignore"
        )
        comparison = comparison.rename(columns={"uber_name": "online_name"}, errors="ignore")
        menus_raw = menus_raw.rename(columns={"uber_name": "online_name"}, errors="ignore")

    # Sort comparison: price mismatches first, then unmatched, then OK
    comp = comparison.copy()
    if "diff" in comp.columns:
        comp["_sort"] = comp["diff"].abs().fillna(999)
    else:
        comp["_sort"] = 999
    comp["_unmatched"] = comp["pos_id"].astype(str).str.len() == 0
    comp = comp.sort_values(["_unmatched", "_sort"], ascending=[True, False]).drop(
        columns=["_sort", "_unmatched"], errors="ignore"
    )
    comp = _order_columns(comp, COMPARISON_COLUMNS)

    # Summary sheet
    matched = comp[comp["pos_id"].astype(str).str.len() > 0]
    with_diff = matched[matched["diff"].notna() & (matched["diff"].abs() > 0.01)] if "diff" in matched.columns else matched.iloc[0:0]
    summary = pd.DataFrame(
        [
            {"Metric": "Uber Eats items", "Value": len(menus_raw)},
            {"Metric": "POS items", "Value": len(pos_raw)},
            {"Metric": "Matched pairs", "Value": len(matched)},
            {"Metric": "Unmatched on Uber", "Value": len(unmatched_online)},
            {"Metric": "Unmatched in POS", "Value": len(unmatched_pos)},
            {"Metric": "Price differences (>£0.01)", "Value": len(with_diff)},
            {"Metric": "Uber higher than POS", "Value": (with_diff["diff"] > 0.01).sum() if len(with_diff) else 0},
            {"Metric": "Uber lower than POS", "Value": (with_diff["diff"] < -0.01).sum() if len(with_diff) else 0},
        ]
    )

    sheets: list[tuple[str, pd.DataFrame, tuple[str, ...]]] = [
        ("Summary", summary, ()),
        ("Comparison", comp, ("online_price", "pos_price", "diff")),
        (
            "Unmatched_Uber",
            _order_columns(unmatched_online, COMPARISON_COLUMNS),
            ("online_price",),
        ),
        ("Unmatched_POS", unmatched_pos, ("pos_price",)),
        ("Uber_Raw", menus_raw, ("online_price",) if "online_price" in menus_raw.columns else ()),
        ("POS_Raw", pos_raw, ("pos_price",)),
    ]

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for name, df, currency_cols in sheets:
            sheet_name = name[:31]
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            _format_sheet(writer.sheets[sheet_name], df, currency_cols=currency_cols)

    return output_path
